from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from douzone_quote_bot.common.files import (
    ensure_parent,
    load_structured_input,
    safe_filename,
)


ROOT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_TEMPLATE = ROOT_DIR / "skills" / "purchase-request" / "templates" / "purchase_request_template.xlsx"


REQUIRED_FIELDS = (
    "customer_name",
    "request_type",
    "product",
    "payment_type",
    "sales_owner",
    "items",
)


def generate_purchase_request(input_path: Path, output_path: Path | None = None) -> Path:
    data = load_structured_input(input_path)
    normalized = _validate(data)

    if output_path is None:
        quote_date = normalized["quote_date"].replace("-", "")
        filename = f"{safe_filename(normalized['customer_name'])}_purchase_request_{quote_date}.xlsx"
        output_path = ROOT_DIR / "outputs" / filename

    ensure_parent(output_path)
    workbook = _load_or_create_workbook()
    _fill_workbook(workbook, normalized)
    workbook.save(output_path)
    return output_path


def _validate(data: dict[str, Any]) -> dict[str, Any]:
    missing = [field for field in REQUIRED_FIELDS if field not in data or data[field] in (None, "")]
    if missing:
        raise ValueError(f"필수 입력값이 없습니다: {', '.join(missing)}")

    items = data["items"]
    if not isinstance(items, list) or not items:
        raise ValueError("items는 1개 이상의 항목을 가진 배열이어야 합니다.")

    normalized_items: list[dict[str, Any]] = []
    for index, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"items[{index}]는 객체여야 합니다.")
        for field in ("name", "quantity", "unit"):
            if field not in item or item[field] in (None, ""):
                raise ValueError(f"items[{index}] 필수 입력값이 없습니다: {field}")
        normalized_items.append(
            {
                "name": str(item["name"]),
                "quantity": item["quantity"],
                "unit": str(item["unit"]),
                "unit_price": item.get("unit_price", "확인 필요"),
                "amount": item.get("amount", "확인 필요"),
                "memo": item.get("memo", ""),
            }
        )

    quote_date = data.get("quote_date") or date.today().isoformat()
    return {
        "customer_name": str(data["customer_name"]),
        "request_type": str(data["request_type"]),
        "product": str(data["product"]),
        "payment_type": str(data["payment_type"]),
        "sales_owner": str(data["sales_owner"]),
        "quote_date": str(quote_date),
        "writer": str(data.get("writer") or data["sales_owner"]),
        "memo": str(data.get("memo") or ""),
        "items": normalized_items,
    }


def _load_or_create_workbook() -> Workbook:
    if DEFAULT_TEMPLATE.exists():
        return load_workbook(DEFAULT_TEMPLATE, data_only=False)
    return _create_fallback_template()


def _create_fallback_template() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "구매신청서"
    worksheet.sheet_view.showGridLines = False

    widths = {
        "A": 6,
        "B": 18,
        "C": 18,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 24,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.merge_cells("A1:G1")
    worksheet["A1"] = "기존고객사 구매신청서"
    worksheet["A1"].font = Font(size=18, bold=True)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 32

    labels = [
        ("A3", "고객사명", "B3"),
        ("D3", "신청구분", "E3"),
        ("A4", "제품", "B4"),
        ("D4", "납부방식", "E4"),
        ("A5", "영업담당", "B5"),
        ("D5", "작성일자", "E5"),
        ("A6", "작성자", "B6"),
    ]
    for label_cell, label, _value_cell in labels:
        worksheet[label_cell] = label

    headers = ["No", "품목", "수량", "단위", "단가", "금액", "비고"]
    for index, header in enumerate(headers, start=1):
        worksheet.cell(row=8, column=index, value=header)

    worksheet.merge_cells("A22:B22")
    worksheet["A22"] = "메모"
    worksheet.merge_cells("C22:G25")

    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    label_fill = PatternFill("solid", fgColor="EAF2F8")
    header_fill = PatternFill("solid", fgColor="D9EAD3")

    for row in worksheet.iter_rows(min_row=3, max_row=6, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for cell_ref in ("A3", "D3", "A4", "D4", "A5", "D5", "A6"):
        worksheet[cell_ref].fill = label_fill
        worksheet[cell_ref].font = Font(bold=True)
        worksheet[cell_ref].alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=8, max_row=20, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for cell in worksheet[8]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=22, max_row=25, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet["A22"].fill = label_fill
    worksheet["A22"].font = Font(bold=True)
    worksheet["A22"].alignment = Alignment(horizontal="center", vertical="center")

    return workbook


def _fill_workbook(workbook: Workbook, data: dict[str, Any]) -> None:
    worksheet = workbook.active
    _set_value(worksheet, "B3", data["customer_name"])
    _set_value(worksheet, "E3", data["request_type"])
    _set_value(worksheet, "B4", data["product"])
    _set_value(worksheet, "E4", data["payment_type"])
    _set_value(worksheet, "B5", data["sales_owner"])
    _set_value(worksheet, "E5", data["quote_date"])
    _set_value(worksheet, "B6", data["writer"])
    _set_value(worksheet, "C22", data["memo"] or "확인 필요")

    start_row = 9
    template_row = start_row
    item_count = len(data["items"])
    if item_count > 12:
        worksheet.insert_rows(start_row + 12, item_count - 12)
        _copy_row_style(worksheet, template_row, start_row + 12, item_count - 12)

    for offset, item in enumerate(data["items"]):
        row = start_row + offset
        values = [
            offset + 1,
            item["name"],
            item["quantity"],
            item["unit"],
            item["unit_price"],
            item["amount"],
            item["memo"],
        ]
        for column, value in enumerate(values, start=1):
            _set_value(worksheet, f"{get_column_letter(column)}{row}", value)


def _set_value(worksheet, cell_ref: str, value: Any) -> None:
    cell = worksheet[cell_ref]
    if cell.coordinate not in worksheet.merged_cells:
        cell.value = value
        return

    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
            return


def _copy_row_style(worksheet, source_row: int, insert_at: int, count: int) -> None:
    for row in range(insert_at, insert_at + count):
        for column in range(1, worksheet.max_column + 1):
            source = worksheet.cell(row=source_row, column=column)
            target = worksheet.cell(row=row, column=column)
            if source.has_style:
                target._style = copy(source._style)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
