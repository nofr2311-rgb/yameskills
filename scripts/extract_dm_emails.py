#!/usr/bin/env python3
"""Extract unique email addresses from raw customer Excel files."""
from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")


def row_text(values: list[object]) -> str:
    return " | ".join(str(v).strip() for v in values if v not in (None, ""))


def extract(paths: list[Path]) -> list[dict[str, object]]:
    seen: set[str] = set()
    rows: list[dict[str, object]] = []
    for path in paths:
        wb = load_workbook(path, data_only=True, read_only=True)
        for ws in wb.worksheets:
            for row_index, values in enumerate(ws.iter_rows(values_only=True), start=1):
                text = row_text(list(values))
                for email in EMAIL_RE.findall(text):
                    key = email.lower()
                    if key in seen:
                        continue
                    seen.add(key)
                    company = first_non_email_value(values)
                    rows.append(
                        {
                            "company": company,
                            "email": email,
                            "copy_email": f"{email},",
                            "source_file": path.name,
                            "sheet": ws.title,
                            "row": row_index,
                            "context": text[:500],
                        }
                    )
    return rows


def first_non_email_value(values: tuple[object, ...]) -> str:
    for value in values:
        if value in (None, ""):
            continue
        text = str(value).strip()
        if EMAIL_RE.fullmatch(text):
            continue
        if len(text) >= 2:
            return text
    return ""


def write(rows: list[dict[str, object]], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "DM_email_list"
    headers = ["상호명", "이메일", "이메일(복사용)", "출처파일", "시트", "행", "원문"]
    ws.append(headers)
    for row in rows:
        ws.append([row["company"], row["email"], row["copy_email"], row["source_file"], row["sheet"], row["row"], row["context"]])

    fill = PatternFill("solid", fgColor="1565C0")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    widths = [30, 35, 38, 30, 20, 8, 80]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, index).column_letter].width = width
    ws.freeze_panes = "A2"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    paths = sorted(list(input_dir.glob("*.xlsx")) + list(input_dir.glob("*.xlsm")))
    if not paths:
        raise SystemExit("엑셀 첨부파일을 찾지 못했습니다.")
    rows = extract(paths)
    write(rows, Path(args.output))
    print(f"files={len(paths)}")
    print(f"emails={len(rows)}")
    print(f"created={args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
