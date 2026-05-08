#!/usr/bin/env python3
"""Generate a WEHAGO new-customer quote xlsx from inbound call text."""
from __future__ import annotations

import argparse
import re
import shutil
import subprocess
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "skills" / "wehago-new-quote" / "templates" / "wehago_template.xlsx"


MODULE_CELLS = {
    "accounting": "M21",
    "hr_payroll": "M22",
    "logistics": "M23",
    "corporate_tax": "M24",
    "personal_tax": "M25",
    "smart_a_2": "M26",
    "smart_a10_usage": "M31",
    "plan_base": "M32",
    "plan_usage": "M33",
    "one_ai": "M41",
    "approval": "M47",
    "attendance": "M48",
    "crm": "M49",
    "pms": "M50",
}

ALIASES = {
    "accounting": ["회계", "재무", "accounting"],
    "hr_payroll": ["인사급여", "인사", "급여", "hr"],
    "logistics": ["물류", "재고"],
    "corporate_tax": ["법인조정", "법인세"],
    "personal_tax": ["개인조정", "개인세"],
    "smart_a_2": ["smart a 2.0", "smarta2", "smart a2"],
    "smart_a10_usage": ["smart a10 사용료", "smart a 10 사용료", "a10 사용료", "사용료"],
    "plan_usage": ["wehago 사용료", "요금제 사용료", "위하고 사용료"],
    "approval": ["전자결재", "결재"],
    "attendance": ["근태", "근태관리"],
    "crm": ["crm"],
    "pms": ["pms"],
    "one_ai": ["one ai", "원에이아이"],
}


def compact(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def find_customer(text: str) -> str:
    patterns = [
        r"고객사\s*[:：]\s*([^\n,]+)",
        r"회사\s*[:：]\s*([^\n,]+)",
        r"상호\s*[:：]\s*([^\n,]+)",
        r"([가-힣A-Za-z0-9㈜(). _-]{2,40})\s*(?:에서|는|은)\s*(?:WEHAGO|위하고|Smart)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if match:
            return match.group(1).strip()
    return "고객사명_확인필요"


def find_qty(text: str, aliases: list[str]) -> int:
    lower = text.lower()
    for alias in aliases:
        alias_re = re.escape(alias.lower()).replace("\\ ", r"\s*")
        patterns = [
            rf"{alias_re}\s*(\d+)\s*(?:인|명|유저|user|ea)?",
            rf"(\d+)\s*(?:인|명|유저|user|ea)?\s*{alias_re}",
        ]
        for pattern in patterns:
            match = re.search(pattern, lower, flags=re.I)
            if match:
                return int(match.group(1))
    return 0


def parse_request(text: str) -> dict[str, object]:
    text = compact(text)
    data: dict[str, object] = {
        "customer": find_customer(text),
        "payment": "연납" if "연납" in text else "월납",
        "plan": "PRO" if "pro" in text.lower() or "프로" in text else "CLUB",
    }
    quantities: dict[str, int] = {}
    for key, aliases in ALIASES.items():
        qty = find_qty(text, aliases)
        if qty:
            quantities[key] = qty

    if "smart_a10_usage" not in quantities:
        base_users = max(quantities.get("accounting", 0), quantities.get("hr_payroll", 0), quantities.get("logistics", 0))
        if base_users:
            quantities["smart_a10_usage"] = base_users
    if "plan_usage" not in quantities and quantities.get("smart_a10_usage"):
        quantities["plan_usage"] = int(quantities["smart_a10_usage"])
    if quantities.get("plan_usage") and "plan_base" not in quantities:
        quantities["plan_base"] = 1

    data["quantities"] = quantities
    return data


def set_cell(ws, cell: str, value) -> None:
    for merged in ws.merged_cells.ranges:
        if cell in merged:
            ws.cell(merged.min_row, merged.min_col).value = value
            return
    ws[cell] = value


def generate(text: str, output: Path, pdf: bool = False) -> Path:
    req = parse_request(text)
    wb = load_workbook(TEMPLATE, data_only=False)
    ws = wb["신규(기업)"]
    set_cell(ws, "G7", req["customer"])
    ws["H29"] = "(연납)" if req["payment"] == "연납" else "(월납)"
    ws["C32"] = f"{req['plan']} 요금제"
    ws["U12"] = "010-4903-2311"
    ws["U13"] = "nofr2311@douzone.com"

    for key, cell in MODULE_CELLS.items():
        ws[cell] = int(req["quantities"].get(key, 0)) if isinstance(req["quantities"], dict) else 0
    for row in range(21, 51):
        if ws[f"R{row}"].value is None:
            set_cell(ws, f"R{row}", 0)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    if pdf:
        export_pdf(output)
    return output


def export_pdf(xlsx: Path) -> None:
    exe = shutil.which("libreoffice") or shutil.which("soffice")
    if not exe:
        print("LibreOffice가 없어 PDF 변환은 건너뜁니다.")
        return
    subprocess.run([exe, "--headless", "--convert-to", "pdf", "--outdir", str(xlsx.parent), str(xlsx)], check=True)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--text", help="Inbound call text")
    parser.add_argument("--text-file", help="Inbound call text file")
    parser.add_argument("--output", required=True)
    parser.add_argument("--pdf", action="store_true")
    args = parser.parse_args()

    if args.text_file:
        text = Path(args.text_file).read_text(encoding="utf-8")
    elif args.text:
        text = args.text
    else:
        raise SystemExit("--text 또는 --text-file이 필요합니다.")
    output = Path(args.output)
    if output.is_dir():
        output = output / f"{date.today():%Y%m%d}_wehago_quote.xlsx"
    generate(text, output, pdf=args.pdf)
    print(f"created={output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
