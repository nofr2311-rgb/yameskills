"""
더존 DM 이메일 리스트 생성 스크립트 (모드 A)

사용법:
    python build_dm_list.py \
        --upgrade <upgrade_file.xlsx> \
        --nsm <nsm_file.xlsx> \
        [--recontract <recontract_file.xlsx>] \
        [--label <담당자명>] \
        --output <output_path.xlsx>
"""

import argparse
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


BATCH_COLORS = ['FFFFFF', 'FFF9C4', 'E8F5E9', 'E3F2FD', 'FCE4EC']
HEADER_COLOR = '1565C0'
THIN = Side(style='thin', color='CCCCCC')
THICK = Side(style='medium', color='888888')


def clean_biz(series):
    return series.str.replace('-', '', regex=False).str.strip()


def load_upgrade(path):
    xl = pd.ExcelFile(path)
    sheet = next((s for s in xl.sheet_names if '위하고교체' in s or '신규' in s), xl.sheet_names[0])
    df = pd.read_excel(path, sheet_name=sheet, dtype=str)
    df['biz'] = clean_biz(df['사업자번호'])
    return df


def load_nsm(path):
    df = pd.read_excel(path, dtype=str)
    df['biz'] = clean_biz(df['사업자번호'])
    return df


def load_recontract(path):
    df = pd.read_excel(path, dtype=str)
    biz_col = '사업자등록번호' if '사업자등록번호' in df.columns else '사업자번호'
    df['biz'] = clean_biz(df[biz_col])
    email_col = '담당자 이메일' if '담당자 이메일' in df.columns else '담당자이메일'
    df['_rc_email'] = df[email_col]
    return df


def build(upgrade_path, nsm_path, recontract_path=None, label='', output_path=None):
    df = load_upgrade(upgrade_path)
    nsm = load_nsm(nsm_path)

    # 이메일 맵 생성
    nsm_map = (nsm.dropna(subset=['담당자이메일'])
                  .drop_duplicates('biz')
                  .set_index('biz')['담당자이메일'])

    rc_map = {}
    if recontract_path:
        rc = load_recontract(recontract_path)
        rc_map = (rc.dropna(subset=['_rc_email'])
                    .drop_duplicates('biz')
                    .set_index('biz')['_rc_email'])

    orig_email = df['이메일'].copy()
    df['_nsm']  = df['biz'].map(nsm_map)
    df['_rc']   = df['biz'].map(rc_map) if rc_map else pd.NA

    df['최종이메일'] = orig_email.fillna(df['_nsm']).fillna(df['_rc'])

    def get_source(i):
        if pd.notna(orig_email.iloc[i]):    return 'WEHAGO'
        if pd.notna(df['_nsm'].iloc[i]):    return 'NSM보완'
        if pd.notna(df.get('_rc', pd.Series()).iloc[i] if '_rc' in df.columns else pd.NA): return '재계약보완'
        return '미확보'

    df['출처'] = [get_source(i) for i in range(len(df))]
    df['D_email'] = df['최종이메일'].apply(
        lambda x: f'{x},' if pd.notna(x) and str(x).strip() not in ['nan', ''] else ''
    )

    # Excel 작성
    wb = Workbook()
    ws = wb.active
    ws.title = 'DM발송대상'

    headers = ['상호명', '사업자번호', '이메일', '이메일 (복사용)', '이메일출처']
    widths  = [30, 16, 35, 35, 10]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        c.fill      = PatternFill('solid', start_color=HEADER_COLOR)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 18

    for i, row in df.iterrows():
        r     = i + 2
        batch = i // 99
        bg    = BATCH_COLORS[batch % len(BATCH_COLORS)]
        fill  = PatternFill('solid', start_color=bg)
        is_end = (i % 99 == 98) or (i == len(df) - 1)
        bot   = THICK if is_end else THIN

        vals = [row.get('거래처', ''), row['사업자번호'],
                row['최종이메일'], row['D_email'], row['출처']]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=v if pd.notna(v) and str(v) != 'nan' else None)
            cell.font      = Font(name='Arial', size=10)
            cell.fill      = fill
            cell.alignment = Alignment(vertical='center')
            cell.border    = Border(bottom=bot, left=THIN, right=THIN)

    # 그룹 라벨
    ws.column_dimensions['G'].width = 14
    for batch in range(10):
        sr  = batch * 99 + 2
        er  = min(sr + 98, len(df) + 1)
        mr  = (sr + er) // 2
        cnt = er - sr + 1
        c   = ws.cell(row=mr, column=7, value=f'그룹{batch+1} ({cnt}개)')
        c.font      = Font(name='Arial', size=9, color='888888', italic=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
        if er >= len(df) + 1:
            break

    ws.freeze_panes = 'A2'

    if not output_path:
        today = date.today().strftime('%Y%m%d')
        prefix = f'{label}_' if label else ''
        output_path = f'{prefix}dm_list_{today}.xlsx'

    wb.save(output_path)

    # 결과 리포트
    src = df['출처'].value_counts().to_dict()
    confirmed = df['최종이메일'].notna().sum()
    missing   = len(df) - confirmed
    groups    = (len(df) - 1) // 99 + 1

    print(f"\n=== {'[' + label + '] ' if label else ''}DM 리스트 결과 ===")
    print(f"전체 대상:     {len(df)}개사")
    print(f"이메일 확보:   {confirmed}개  "
          f"(WEHAGO: {src.get('WEHAGO',0)} / NSM보완: {src.get('NSM보완',0)} / "
          f"재계약보완: {src.get('재계약보완',0)})")
    print(f"이메일 미확보: {missing}개")
    print(f"99개 그룹:    {groups}그룹")
    if missing:
        print(f"\n⚠️  미확보 {missing}개는 수동 확인이 필요합니다.")
    print(f"\n저장 완료: {output_path}")

    return output_path


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--upgrade',    required=True)
    parser.add_argument('--nsm',        required=True)
    parser.add_argument('--recontract', default=None)
    parser.add_argument('--label',      default='')
    parser.add_argument('--output',     default=None)
    args = parser.parse_args()

    build(args.upgrade, args.nsm, args.recontract, args.label, args.output)
