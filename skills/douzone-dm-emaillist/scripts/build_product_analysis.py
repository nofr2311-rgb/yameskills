"""
더존 고객 구성 비율 + 솔루션 이메일 추출 스크립트 (모드 B)

사용법:
    python build_product_analysis.py \
        --nsm <nsm_file.xlsx> \
        --output <output_path.xlsx>
"""

import argparse
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference


HEADER_COLOR = '1565C0'
THIN = Side(style='thin', color='CCCCCC')

PROD_RULES = {
    'WEHAGO (플랫폼)': 'WEHAGO',
    'ICUBE':           'ICUBE',
    'Bizbox Alpha':    'Bizbox',
    'Amaranth 10':     'Amaranth 10',
}

ROW_COLORS = {
    'WEHAGO (플랫폼)': 'E3F2FD',
    'ICUBE':           'E8F5E9',
    'Bizbox Alpha':    'FFF9C4',
    'Amaranth 10':     'FCE4EC',
    '기타':            'F5F5F5',
    '전체':            'BBDEFB',
}

SOL_COLORS = {
    'ICUBE':       'E8F5E9',
    'Bizbox Alpha':'FFF9C4',
    'Amaranth 10': 'FCE4EC',
}


def build(nsm_path, output_path=None):
    df = pd.read_excel(nsm_path, dtype=str)
    prod = df['구매 제품군'].fillna('')
    total = len(df)

    masks = {k: prod.str.contains(v, na=False) for k, v in PROD_RULES.items()}
    sol_mask = masks['ICUBE'] | masks['Bizbox Alpha'] | masks['Amaranth 10']

    # 기타 계산 (WEHAGO / 솔루션 모두 없는 경우)
    other_mask = ~masks['WEHAGO (플랫폼)'] & ~sol_mask

    rows = []
    for name, mask in masks.items():
        rows.append((name, mask.sum(), round(mask.sum() / total * 100, 1)))
    rows.append(('기타 (Smart A 등)', other_mask.sum(), round(other_mask.sum() / total * 100, 1)))
    rows.append(('전체', total, 100.0))

    # ===== Excel 작성 =====
    wb = Workbook()

    # ---- Sheet 1: 고객구성비율 ----
    ws1 = wb.active
    ws1.title = '고객구성비율'

    for ci, (h, w) in enumerate(zip(['제품', '고객사 수', '비율(%)', '비고'],
                                     [22, 10, 10, 36]), 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        c.fill      = PatternFill('solid', start_color=HEADER_COLOR)
        c.alignment = Alignment(horizontal='center')
        ws1.column_dimensions[c.column_letter].width = w

    NOTES = {
        'WEHAGO (플랫폼)': '위하고/위하고T 계약 고객사',
        'ICUBE':           '솔루션',
        'Bizbox Alpha':    '솔루션',
        'Amaranth 10':     '솔루션',
        '기타 (Smart A 등)': 'Smart A / ERP-iU 등',
        '전체':            '',
    }

    for ri, (name, cnt, pct) in enumerate(rows, 2):
        bg   = ROW_COLORS.get(name.split('(')[0].strip(), 'FFFFFF')
        fill = PatternFill('solid', start_color=bg)
        bold = (name == '전체')
        for ci, v in enumerate([name, cnt, pct, NOTES.get(name, '')], 1):
            cell = ws1.cell(row=ri, column=ci, value=v)
            cell.font      = Font(name='Arial', size=10, bold=bold)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal='center' if ci in [2, 3] else 'left',
                                       vertical='center')
            cell.border    = Border(bottom=THIN, left=THIN, right=THIN)

    # 파이차트 (기타/전체 제외, 1~4번 제품)
    pie = PieChart()
    pie.title  = '고객 구성 비율'
    pie.style  = 10
    pie.width  = 14
    pie.height = 10
    labels = Reference(ws1, min_col=1, min_row=2, max_row=5)
    data   = Reference(ws1, min_col=2, min_row=1, max_row=5)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws1.add_chart(pie, 'F2')

    # ---- Sheet 2: 솔루션 이메일 ----
    ws2 = wb.create_sheet('솔루션고객사_이메일')
    sol_df = df[sol_mask].copy()

    def get_prod_str(i):
        parts = []
        for name, kw in [('ICUBE','ICUBE'),('Bizbox Alpha','Bizbox'),('Amaranth 10','Amaranth 10')]:
            if kw in str(sol_df.iloc[i]['구매 제품군']):
                parts.append(name)
        return ' / '.join(parts)

    sol_df = sol_df.reset_index(drop=True)
    sol_df['제품구성']   = [get_prod_str(i) for i in range(len(sol_df))]
    sol_df['WEHAGO여부'] = sol_df.index.map(
        lambda i: 'O' if 'WEHAGO' in str(sol_df.iloc[i]['구매 제품군']) else ''
    )
    sol_df['이메일']     = sol_df['담당자이메일']
    sol_df['D_email']    = sol_df['이메일'].apply(
        lambda x: f'{x},' if pd.notna(x) and str(x).strip() not in ['nan', ''] else ''
    )

    headers2 = ['상호명', '사업자번호', '제품구성', 'WEHAGO여부', '이메일', '이메일 (복사용)']
    widths2  = [30, 16, 28, 12, 35, 35]

    for ci, (h, w) in enumerate(zip(headers2, widths2), 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        c.fill      = PatternFill('solid', start_color=HEADER_COLOR)
        c.alignment = Alignment(horizontal='center')
        ws2.column_dimensions[c.column_letter].width = w

    for i, row in sol_df.iterrows():
        r   = i + 2
        p0  = row['제품구성'].split(' / ')[0] if row['제품구성'] else ''
        bg  = SOL_COLORS.get(p0, 'F3E5F5')
        fill = PatternFill('solid', start_color=bg)
        vals = [row.get('거래처명', row.get('거래처', '')), row['사업자번호'],
                row['제품구성'], row['WEHAGO여부'], row['이메일'], row['D_email']]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=ci, value=v if pd.notna(v) and str(v) != 'nan' else None)
            cell.font      = Font(name='Arial', size=10)
            cell.fill      = fill
            cell.alignment = Alignment(vertical='center')
            cell.border    = Border(bottom=THIN, left=THIN, right=THIN)

    ws1.freeze_panes = 'A2'
    ws2.freeze_panes = 'A2'

    if not output_path:
        today = date.today().strftime('%Y%m%d')
        output_path = f'고객구성비율_솔루션이메일_{today}.xlsx'

    wb.save(output_path)

    # 리포트
    print('\n=== 고객 구성 비율 분석 결과 ===')
    for name, cnt, pct in rows:
        bar = '█' * int(pct / 2)
        print(f'{name:<20} {cnt:>5}개사  {pct:>5.1f}%  {bar}')
    print(f'\n솔루션 고객사: {sol_mask.sum()}개사  '
          f'(이메일 확보: {sol_df["이메일"].notna().sum()} / {len(sol_df)})')
    print(f'\n저장 완료: {output_path}')

    return output_path


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--nsm',    required=True)
    parser.add_argument('--output', default=None)
    args = parser.parse_args()

    build(args.nsm, args.output)
