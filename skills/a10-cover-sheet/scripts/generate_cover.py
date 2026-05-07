#!/usr/bin/env python3
"""
Amaranth 10 갑지(가격제안서 요약) 생성기
입력: xlsx(갑지 시트 포함) 또는 pdf(갑지 PDF)
출력: pdf + xlsx 모두 생성

v2: 데이터 없는 유형 열 숨김 + 값 없는 항목 행 숨김

Usage: python3 generate_cover.py <input.xlsx|pdf> <output_basename>
"""
import sys, os, subprocess, re, shutil
from datetime import datetime, timedelta
from copy import copy
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent.parent
TEMPLATE_PATH = SCRIPT_DIR / "assets" / "galji_template.xlsx"
COLS = ["J","N","R","V"]

# 각 유형별 열 범위 (헤더 포함)
COL_GROUPS = {
    0: ["J","K","L","M"],      # Cloud형
    1: ["N","O","P","Q"],      # 구축형(서버 파킹)
    2: ["R","S","T","U"],      # 구축형(임대)
    3: ["V","W","X","Y"],      # 구축형(자체)
}

# 절대 숨기지 않는 행 (섹션 헤더)
NEVER_HIDE_ROWS = {19}  # EBP 엔진 (Enterprise Business Platform 섹션 헤더)

# 기본 제안사 정보
DEFAULT_PROVIDER = "134-81-08473 (주)더존비즈온\n강원도 춘천시 남산면 버들1길 130\n대표이사 이 강 수, 지 용 구 (직인생략)"
DEFAULT_SALES_REP = "이영진 대리"
DEFAULT_SALES_CONTACT = "010-4903-2311"

COST_ROWS = [
    (19,"ebp_engine","Enterprise\nBusiness Platform","엔진"),
    (21,"erp_module","","ERP 모듈"),
    (22,"uc_module","","UC(그룹웨어) 모듈"),
    (23,"user_license","","유저 라이선스"),
    (24,"approval_mig","","전자결재 마이그레이션"),
    (25,"board_mig","","게시판 마이그레이션"),
    (26,"education","","구축 교육비"),
    (27,"sw_license","","필수 S/W 라이선스"),
    (28,"hw_purchase","","서버 H/W 매입"),
]
SERVICE_ROWS = [
    (29,"monthly_svc","","월납 비용(사용료 등)"),
    (30,"monthly_oneai","","월납 비용(ONE AI)"),
    (31,"annual_maint","","연납 비용(유지보수)"),
]
TOTAL_ROWS = [
    (32,"total_onetime","일회납 비용","라이선스 및 구축비"),
    (33,"total_monthly","월납 비용","SSL인증서, IDC사용, ONE AI"),
    (34,"total_annual","연납 비용","유지보수비용"),
]

def fmt_money(val):
    if val is None or val=="" or val==0: return "-"
    if isinstance(val,str):
        try: n=int(val.replace(",","")); return f"{n:,}" if n else "-"
        except: return val
    try: n=int(val); return f"{n:,}" if n else "-"
    except: return str(val)

def fmt_date(val):
    if isinstance(val,datetime): return f"{val.year}년 {val.month}월 {val.day}일"
    if isinstance(val,(int,float)):
        d=datetime(1899,12,30)+timedelta(days=int(val))
        return f"{d.year}년 {d.month}월 {d.day}일"
    return str(val) if val else ""

def safe_str(val): return str(val) if val is not None else ""


def _is_empty_val(v):
    """값이 비어있는지 판별"""
    if v is None or v == "" or v == 0 or v == "-":
        return True
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        return s in ("", "0", "-")
    try:
        return int(v) == 0
    except:
        return False


def detect_active_columns(data):
    """총계 행 + 개별 항목 기준으로 데이터 있는 열 인덱스 리스트 반환"""
    active = set()
    for _, key, _, _ in TOTAL_ROWS + COST_ROWS + SERVICE_ROWS:
        vals = data.get(key, [0, 0, 0, 0])
        for i in range(4):
            v = vals[i] if i < len(vals) else 0
            if not _is_empty_val(v):
                active.add(i)
    return sorted(active)


def detect_empty_rows(data, active_cols):
    """활성 열 기준으로 모든 값이 비어있는 항목 행 번호 리스트 반환 (총계 행 제외)"""
    empty_rows = []
    for row_num, key, _, _ in COST_ROWS + SERVICE_ROWS:
        if row_num in NEVER_HIDE_ROWS:
            continue
        vals = data.get(key, [0, 0, 0, 0])
        all_empty = True
        for i in active_cols:
            v = vals[i] if i < len(vals) else 0
            if not _is_empty_val(v):
                all_empty = False
                break
        if all_empty:
            empty_rows.append(row_num)
    return empty_rows


# ── XLSX 입력 → 데이터 추출 ──
def read_from_xlsx(path):
    import openpyxl
    wb=openpyxl.load_workbook(path,data_only=True)
    if "갑지" not in wb.sheetnames: raise ValueError("'갑지' 시트 없음")
    ws=wb["갑지"]; d={}
    for k,ref in {"company":"F5","contact":"F6","title":"F7","quote_no":"F9",
                   "date":"W5","provider":"W6","sales_rep":"W8","sales_contact":"W9"}.items():
        d[k]=ws[ref].value
    d["date_fmt"]=fmt_date(d["date"])
    d["server_purchase"]=[ws[f"{c}13"].value for c in COLS]
    d["infra"]=[ws[f"{c}14"].value for c in COLS]
    d["erp_modules"]=safe_str(ws["J15"].value)
    d["uc_modules"]=safe_str(ws["J16"].value)
    d["user_config"]=safe_str(ws["J17"].value)
    for row,key,_,_ in COST_ROWS+SERVICE_ROWS+TOTAL_ROWS:
        d[key]=[ws[f"{c}{row}"].value for c in COLS]
    d["remarks"]=[str(ws[f"B{r}"].value) for r in range(37,42) if ws[f"B{r}"].value]
    wb.close(); return d


# ── PDF 입력 → 데이터 추출 ──
def read_from_pdf(path):
    import pdfplumber
    d={"company":"","contact":"","title":"Amaranth 10 견적","quote_no":"",
       "date_fmt":"","provider":"","sales_rep":"","sales_contact":"","remarks":[]}
    with pdfplumber.open(path) as pdf:
        text=pdf.pages[0].extract_text()
        tables=pdf.pages[0].extract_tables()
    for line in text.split("\n"):
        if "업체명" in line:
            # "업체명 페더럴익스프레스코리아 유한회사 견적일자 ..." → 견적일자 앞까지 추출
            after = line.split("업체명")[-1].strip()
            if "견적일자" in after:
                d["company"] = after.split("견적일자")[0].strip()
            else:
                d["company"] = after.strip()
        m=re.search(r'(\d{4})년\s*(\d+)월\s*(\d+)일',line)
        if m and not d["date_fmt"]: d["date_fmt"]=f"{m.group(1)}년 {m.group(2)}월 {m.group(3)}일"
        if "견적번호" in line:
            m2=re.search(r'(Douzone-A10-[\w.\-]+)',line)
            if m2: d["quote_no"]=m2.group(1)
        if line.startswith(("1.","2.","3.","4.","5.","※")): d["remarks"].append(line)
    for _,key,_,_ in COST_ROWS+SERVICE_ROWS+TOTAL_ROWS: d[key]=[0,0,0,0]
    d["server_purchase"]=["","","",""]
    d["infra"]=["","","",""]
    d["erp_modules"]=d["uc_modules"]=d["user_config"]=""

    # 제안 구성 테이블에서 추출 (서버 구매, 인프라, 모듈, 사용자)
    for tbl in tables:
        for row in tbl:
            if not row: continue
            rtxt = " ".join(safe_str(c) for c in row)
            if "서버 구매" in rtxt:
                # columns: [None, '서버 구매', Cloud, 파킹, 임대, 자체, 비고]
                vals = row[2:6] if len(row) >= 6 else row[2:]
                d["server_purchase"] = [safe_str(v) for v in vals] + [""]*(4-len(vals))
            elif "인프라 구성" in rtxt:
                vals = row[2:6] if len(row) >= 6 else row[2:]
                d["infra"] = [safe_str(v) for v in vals] + [""]*(4-len(vals))
            elif "ERP 모듈 구성" in rtxt:
                d["erp_modules"] = safe_str(row[2]) if len(row) > 2 and row[2] else ""
            elif "UC(그룹웨어) 모듈 구성" in rtxt or "UC 모듈 구성" in rtxt:
                d["uc_modules"] = safe_str(row[2]) if len(row) > 2 and row[2] else ""
            elif "사용자 구성" in rtxt:
                d["user_config"] = safe_str(row[2]) if len(row) > 2 and row[2] else ""

    if tables:
        label_map={"ERP 모듈":"erp_module","UC(그룹웨어) 모듈":"uc_module",
            "유저 라이선스":"user_license","전자결재 마이그레이션":"approval_mig",
            "게시판 마이그레이션":"board_mig","구축 교육비":"education",
            "필수 S/W 라이선스":"sw_license","서버 H/W 매입":"hw_purchase",
            "월납 비용(사용료 등)":"monthly_svc","월납 비용(ONE AI)":"monthly_oneai",
            "연납 비용(유지보수)":"annual_maint","일회납 비용":"total_onetime",
            "월납 비용":"total_monthly","연납 비용":"total_annual"}

        def _parse_cell_val(cell):
            """셀 값을 숫자 또는 0으로 변환"""
            s = safe_str(cell).strip().replace(",","")
            if not s or s == "-" or s == "None":
                return 0
            try:
                return int(s)
            except:
                return 0

        # 비용 테이블 찾기 (가장 긴 테이블)
        tbl = max(tables, key=len)

        # 헤더 행에서 Cloud/파킹/임대/자체 열 인덱스 찾기
        val_col_start = None
        for row in tbl:
            if not row: continue
            rtxt = " ".join(safe_str(c) for c in row)
            # 헤더 행 감지: "엔진" 또는 "Enterprise" 포함
            if "엔진" in rtxt or "Enterprise" in rtxt:
                # 첫 번째 데이터 행 — 값 열은 라벨 뒤의 4개 열
                # 라벨 열 수 파악 (None이 아닌 텍스트가 끝나는 지점)
                for ci, cell in enumerate(row):
                    s = safe_str(cell).strip().replace(",","")
                    if s and (s == "-" or s.replace(",","").isdigit()):
                        val_col_start = ci
                        break
                break

        if val_col_start is None:
            val_col_start = 3  # fallback

        for row in tbl:
            if not row: continue
            rtxt = " ".join(safe_str(c) for c in row)
            for lbl, key in label_map.items():
                if lbl in rtxt:
                    # 값 열 4개 추출 (val_col_start부터)
                    vals = [0, 0, 0, 0]
                    for i in range(4):
                        ci = val_col_start + i
                        if ci < len(row):
                            vals[i] = _parse_cell_val(row[ci])
                    d[key] = vals
                    break
    return d


# ── XLSX 출력 (템플릿 기반 + 열/행 숨김) ──
def write_xlsx(data, output):
    import openpyxl
    from openpyxl.cell.cell import MergedCell

    active_cols = detect_active_columns(data)
    empty_rows = detect_empty_rows(data, active_cols)
    inactive_cols = [i for i in range(4) if i not in active_cols]

    print(f"  활성 유형: {[['Cloud','파킹','임대','자체'][i] for i in active_cols]}")
    if inactive_cols:
        print(f"  숨김 유형: {[['Cloud','파킹','임대','자체'][i] for i in inactive_cols]}")
    if empty_rows:
        print(f"  숨김 행: {empty_rows}")

    if TEMPLATE_PATH.exists():
        wb=openpyxl.load_workbook(str(TEMPLATE_PATH)); ws=wb["갑지"]
        mrs=list(ws.merged_cells.ranges)
        for mr in mrs: ws.unmerge_cells(str(mr))

        # 데이터 쓰기
        ws["F5"]=data.get("company","")
        ws["F6"]=data.get("contact","")
        ws["F7"]=data.get("title","")
        ws["F9"]=""  # 견적번호 생략
        if isinstance(data.get("date"),datetime): ws["W5"]=data["date"]
        else: ws["W5"]=data.get("date_fmt","")
        ws["W6"]=data.get("provider", DEFAULT_PROVIDER).replace("<br>","\n") or DEFAULT_PROVIDER
        ws["W8"]=data.get("sales_rep","") or DEFAULT_SALES_REP
        ws["W9"]=data.get("sales_contact","") or DEFAULT_SALES_CONTACT
        for key,row in [("server_purchase",13),("infra",14)]:
            vals=data.get(key,["","","",""])
            for i,c in enumerate(COLS):
                ws[f"{c}{row}"]=vals[i] if i<len(vals) else ""
        ws["J15"]=data.get("erp_modules","")
        ws["J16"]=data.get("uc_modules","")
        ws["J17"]=data.get("user_config","")
        for row,key,_,_ in COST_ROWS+SERVICE_ROWS+TOTAL_ROWS:
            vals=data.get(key,[0,0,0,0])
            for i,c in enumerate(COLS):
                ws[f"{c}{row}"]=vals[i] if i<len(vals) else 0
        remarks=data.get("remarks",[])
        for idx,r in enumerate(range(37,42)):
            ws[f"B{r}"]=remarks[idx] if idx<len(remarks) else ""

        # ── 비활성 열 숨김 ──
        for col_idx in inactive_cols:
            for col_letter in COL_GROUPS[col_idx]:
                ws.column_dimensions[col_letter].hidden = True

        # ── 빈 항목 행 숨김 ──
        for row_num in empty_rows:
            ws.row_dimensions[row_num].hidden = True

        # 병합 복원
        for mr in mrs:
            ws.merge_cells(str(mr))

        # ── 인쇄 설정: 가운데 정렬 + 최소 여백 ──
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.3
        ws.page_margins.bottom = 0.3
        ws.page_margins.header = 0.15
        ws.page_margins.footer = 0.15
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True
    else:
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="갑지"
        ws["B1"]="Amaranth 10 가격제안서 요약"
        r=3
        for _,key,_,lbl in COST_ROWS+SERVICE_ROWS+TOTAL_ROWS:
            vals=data.get(key,[0,0,0,0])
            ws[f"B{r}"]=lbl
            for i,c in enumerate(["D","F","H","J"]):
                ws[f"{c}{r}"]=vals[i] if i<len(vals) else 0
            r+=1
    wb.save(output)
    print(f"✅ XLSX: {output}")


# ── PDF 출력 (숨김 적용된 xlsx → LibreOffice 렌더링) ──
def write_pdf_from_xlsx(xlsx_path, output):
    """
    숨김 적용된 갑지 xlsx를 LibreOffice로 PDF 변환 후 갑지 페이지만 추출.
    """
    from pypdf import PdfReader, PdfWriter

    tmp_xlsx = output.replace(".pdf", "_tmp_src.xlsx")
    shutil.copy2(xlsx_path, tmp_xlsx)

    outdir = os.path.dirname(output) or "."
    tmp_pdf = os.path.join(outdir, Path(tmp_xlsx).stem + ".pdf")

    subprocess.run([
        "libreoffice", "--headless", "--calc",
        "--convert-to", "pdf",
        "--outdir", outdir,
        tmp_xlsx
    ], capture_output=True, text=True, timeout=60)

    if not os.path.exists(tmp_pdf):
        print("❌ LibreOffice PDF 변환 실패")
        return

    reader = PdfReader(tmp_pdf)
    galji_page_idx = None

    for i, page in enumerate(reader.pages):
        txt = page.extract_text()[:200] if page.extract_text() else ""
        if "하기와 같이 견적합니다" in txt:
            galji_page_idx = i
            break
        if "업체명" in txt and "견적일자" in txt and "구분" in txt:
            galji_page_idx = i
            break

    if galji_page_idx is None:
        galji_page_idx = 1 if len(reader.pages) > 1 else 0
        print(f"⚠️ 갑지 페이지 자동탐색 실패, page {galji_page_idx} 사용")

    writer = PdfWriter()
    writer.add_page(reader.pages[galji_page_idx])
    with open(output, "wb") as f:
        writer.write(f)

    try:
        os.remove(tmp_xlsx)
        os.remove(tmp_pdf)
    except: pass

    if os.path.exists(output):
        print(f"✅ PDF: {output} (page {galji_page_idx} from LibreOffice)")
    else:
        print("❌ PDF 추출 실패")


# ── main ──
def main():
    if len(sys.argv)<3:
        print("Usage: python3 generate_cover.py <input.xlsx|pdf> <output_basename>")
        sys.exit(1)
    inp=sys.argv[1]; base=sys.argv[2]
    if base.endswith((".pdf",".xlsx")): base=base.rsplit(".",1)[0]
    if not os.path.exists(inp): print(f"Error: {inp} not found"); sys.exit(1)
    ext=Path(inp).suffix.lower()

    print(f"입력: {ext}")
    if ext==".xlsx":
        data=read_from_xlsx(inp)

        # 1) XLSX 먼저 생성 (열/행 숨김 적용)
        xlsx_out = f"{base}.xlsx"
        print("XLSX 생성 (열/행 숨김 적용)...")
        write_xlsx(data, xlsx_out)

        # 2) 숨김 적용된 XLSX로 PDF 생성
        print("PDF 생성 (LibreOffice 렌더링)...")
        write_pdf_from_xlsx(xlsx_out, f"{base}.pdf")

    elif ext==".pdf":
        data=read_from_pdf(inp)

        # XLSX 생성 (열/행 숨김 적용)
        xlsx_out = f"{base}.xlsx"
        print("XLSX 생성 (열/행 숨김 적용)...")
        write_xlsx(data, xlsx_out)

        # 숨김 적용된 XLSX로 PDF 생성
        print("PDF 생성 (LibreOffice 렌더링)...")
        write_pdf_from_xlsx(xlsx_out, f"{base}.pdf")

    else:
        print(f"지원 안됨: {ext}"); sys.exit(1)

    if "date_fmt" not in data or not data["date_fmt"]:
        data["date_fmt"]=fmt_date(data.get("date",""))

    print(f"\n📁 출력: {base}.pdf + {base}.xlsx")

if __name__=="__main__": main()
